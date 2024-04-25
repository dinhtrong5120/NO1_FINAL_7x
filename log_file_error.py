import pandas as pd
import os
from check_input_v2_for_new_function import create_status_all_karen_file
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import unicodedata


def normalize_japanese_text(input_text):
    normalized_text = ''
    if isinstance(input_text, str):
        for char in input_text:
            normalized_char = unicodedata.normalize('NFKC', char)
            normalized_text += normalized_char
        normalized_text = normalized_text.replace("\n", "")
        normalized_text = normalized_text.strip()
        return normalized_text
    else:
        return input_text


def check_document(link_folder, file_spec,folder_out):
    karen1_files = [f for f in os.listdir(link_folder) if f.endswith('.xlsx') and f.startswith('関連表1')]
    karen2_files = [f for f in os.listdir(link_folder) if f.endswith('.xlsx') and f.startswith('関連表2')]
    karen3_files = [f for f in os.listdir(link_folder) if f.endswith('.xlsx') and f.startswith('関連表3')]
    karen4_files = [f for f in os.listdir(link_folder) if f.endswith('.xlsx') and f.startswith('関連表4')]
    df_karen1 = check_karen_1(karen1_files, link_folder)
    df_karen2 = check_karen_2(karen2_files, link_folder, file_spec)
    df_karen3 = check_karen_3(karen3_files, link_folder)
    df_karen4 = check_karen_4(karen4_files, link_folder)
    df_supper = create_status_all_karen_file(karen1_files, karen2_files, karen3_files, karen4_files)
    excel_file = folder_out+"/"+'File Log.xlsx'
    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
        df_supper.to_excel(writer, sheet_name='COUNT', index=False)
        df_karen1.to_excel(writer, sheet_name='関連表①', index=False)
        df_karen2.to_excel(writer, sheet_name='関連表②', index=False)
        df_karen3.to_excel(writer, sheet_name='関連表③', index=False)
        df_karen4.to_excel(writer, sheet_name='関連表④', index=False)
    wb = openpyxl.load_workbook(excel_file)
    ws1 = wb['COUNT']
    fill_pattern_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    fill_pattern_yellow = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    max_columns_in_wb_output = ws1.max_column
    skip_first_row = True
    for row in ws1.iter_rows():
        for cell in row:
            if skip_first_row:
                skip_first_row = False
                ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1, value='Comment')
                cell_comment = ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1)
                cell_comment.font = Font(bold=True)
                break
            if cell.value is None:
                cell.fill = fill_pattern_gray
            for cell in row:
                if cell.value is not None:
                    if ') ' in cell.value:
                        cell.fill = fill_pattern_yellow
                        ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1, value='”)” 後にSpaceがあり')
                    if cell.value[6] == "_":
                        cell.fill = fill_pattern_yellow
                        ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1, value='”_”前にSpaceがあり')
                    elif cell.value[5] == "(":
                        cell.fill = fill_pattern_yellow
                        ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1, value='”A, B, C,...”情報に足りない')
                    elif cell.value.count('_') == 1:
                        cell.fill = fill_pattern_yellow
                        ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1, value='Gr情報に足りない')
                    elif ' .xlsx' in cell.value:
                        cell.fill = fill_pattern_yellow
                        ws1.cell(row=cell.row, column=max_columns_in_wb_output + 1, value='後にSpaceがあり')

    for column in ws1.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 4)
        ws1.column_dimensions[column[0].column_letter].width = adjusted_width
    for item in ['関連表①', '関連表②', '関連表③', '関連表④']:
        ws = wb[item]
        for row in range(2, ws.max_row + 1):
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        max_length = 0
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
            for cell in row:
                max_length = max(max_length, len(str(cell.value))) + 0.7

        # Đặt độ rộng cho cột A
        ws.column_dimensions['A'].width = max_length
    wb.save(excel_file)
    if df_karen1.empty and df_karen2.empty and df_karen3.empty and df_karen4.empty:
        return "Completed!"
    else:
        message = "Completed!, Check File Log _ Sheet: "
        if not df_karen1.empty:
            message += "\n関連表①"
        if not df_karen2.empty:
            message += "\n関連表②"
        if not df_karen3.empty:
            message += "\n関連表③"
        if not df_karen4.empty:
            message += "\n関連表④"
        return message


def check_karen_1(karen1_files, karen1_path):
    list_sheet_name = ['パターン', '関連表']
    list_end = []
    for file_name in karen1_files:
        file_path = karen1_path + '/' + file_name
        all_sheets = pd.ExcelFile(file_path).sheet_names
        dict_sub = {'file_name': None, 'Sheet パターン': None, 'Sheet 関連表': None, 'EV': None, 'e-Power': None,
                    'ICE': None, 'CADICS No.': None, 'パターンNo.': None}
        for item in list_sheet_name:
            if item in all_sheets:
                if item == 'パターン':
                    df = pd.read_excel(file_path, sheet_name=item, header=None)
                    if df.iloc[0, 0] != 'EV':
                        dict_sub['file_name'] = file_name
                        dict_sub['EV'] = '✕'
                    if df.iloc[0, 10] != 'e-Power':
                        dict_sub['file_name'] = file_name
                        dict_sub['e-Power'] = f'✕'
                    if df.iloc[0, 20] != 'ICE':
                        dict_sub['file_name'] = file_name
                        dict_sub['ICE'] = '✕'
                elif item == '関連表':
                    df = pd.read_excel(file_path, sheet_name=item, header=None)
                    if df.iloc[6, 0] != 'CADICS No.':
                        dict_sub['file_name'] = file_name
                        dict_sub['CADICS No.'] = '✕'
                    if df.iloc[1, 10] != 'パターンNo.':
                        dict_sub['file_name'] = file_name
                        dict_sub['パターンNo.'] = '✕'
            else:
                dict_sub['file_name'] = file_name
                dict_sub[f'Sheet {item}'] = '✕'
        if list(set(dict_sub.values())) != [None]:
            list_end.append(dict_sub)
    df = pd.DataFrame(list_end)
    return df


def check_karen_2(karen2_files, karen2_path, file_spec_f):
    list_end = []
    for file_name in karen2_files:
        file_path = karen2_path + '/' + file_name
        all_sheets = pd.ExcelFile(file_path).sheet_names
        dict_sub = {'file_name': None, 'Sheet 関連表': None, 'CADICS No.': None, 'SDN': None, 'H/B': None, 'SUV': None,
                    'MiniVAN': None, 'FRAME': None, 'Zone Region': None, 'Key Error': None}
        if '関連表' in all_sheets:
            df = pd.read_excel(file_path, sheet_name='関連表', header=None)
            if df.iloc[4, 0] != 'CADICS No.':
                dict_sub['file_name'] = file_name
                dict_sub['CADICS No.'] = '✕'

            df_1 = df.map(lambda x: normalize_japanese_text(x).lower() if isinstance(x, str) else x)
            flag_check_empty = False

            matching_columns_hb = df_1.columns[df_1.iloc[1].apply(lambda x: str(x).lower()) == 'h/b'].tolist()
            if len(matching_columns_hb) == 0:
                dict_sub['file_name'] = file_name
                dict_sub['H/B'] = '✕'

            matching_columns_sdn = df_1.columns[df_1.iloc[1].apply(lambda x: str(x).lower()) == 'sdn'].tolist()
            if len(matching_columns_sdn) == 0:
                dict_sub['file_name'] = file_name
                dict_sub['SDN'] = '✕'

            matching_columns_suv = df_1.columns[df_1.iloc[1].apply(lambda x: str(x).lower()) == 'suv'].tolist()
            if len(matching_columns_suv) == 0:
                dict_sub['file_name'] = file_name
                dict_sub['SUV'] = '✕'

            matching_columns_minivan = df_1.columns[df_1.iloc[1].apply(lambda x: str(x).lower()) == 'minivan'].tolist()
            if len(matching_columns_minivan) == 0:
                dict_sub['file_name'] = file_name
                dict_sub['MiniVAN'] = '✕'

            matching_columns_frame = df_1.columns[df_1.iloc[1].apply(lambda x: str(x).lower()) == 'frame'].tolist()
            if len(matching_columns_frame) == 0:
                dict_sub['file_name'] = file_name
                dict_sub['FRAME'] = '✕'

            matching_columns_zone = df_1.columns[df_1.iloc[1].apply(lambda x: str(x).lower()) == 'zone'].tolist()
            if len(matching_columns_zone) == 0:
                dict_sub['file_name'] = file_name
                dict_sub['ZONE'] = '✕'
                flag_check_empty = True
                df_temp = pd.DataFrame()
            else:
                data_test = df_1.iloc[1:3, max(matching_columns_zone) + 1:]
                if data_test.empty or data_test.isna().all().all():
                    flag_check_empty = True
                    df_temp = data_test
                else:
                    flag_check_empty = False
                    data_test.iloc[0] = data_test.iloc[0].str.strip()
                    df_temp = data_test.reset_index(drop=True)
                if not flag_check_empty:
                    df_x = df_temp.iloc[0, :]
                    df = df_x.drop_duplicates()
                    df = pd.DataFrame(df).reset_index(drop=True).dropna()
                    if os.path.exists(file_spec_f):
                        data_spec_ = pd.read_excel(file_spec_f, sheet_name="Sheet1", header=None)
                        data_spec_ = data_spec_.map(
                            lambda x: normalize_japanese_text(x).lower() if isinstance(x, str) else x)
                        data_spec_.iloc[:, 3] = data_spec_.iloc[:, 3].str.strip()
                        for index, value in df[0].items():
                            rows_with_value = data_spec_.index[data_spec_[3] == value.strip()]
                            if rows_with_value.empty:
                                dict_sub['file_name'] = file_name
                                if dict_sub['Key Error'] is None:
                                    dict_sub['Key Error'] = value
                                else:
                                    x = ', ' + value
                                    dict_sub['Key Error'] += x

        else:
            dict_sub['file_name'] = file_name
            dict_sub[f'Sheet 関連表'] = '✕'
        if list(set(dict_sub.values())) != [None]:
            list_end.append(dict_sub)
    df = pd.DataFrame(list_end)
    return df


def check_karen_3(karen3_files, karen3_path):
    list_sheets = ['関連表VC', '関連表PT1', '関連表PT2', '関連表PFC']
    list_end = []
    for file_name in karen3_files:
        file_path = karen3_path + '/' + file_name
        all_sheets = pd.ExcelFile(file_path).sheet_names
        dict_sub = {'file_name': None, '関連表VC_': None, '関連表PT1_': None, '関連表PT2_': None,
                    '関連表PFC_': None,
                    'CADICS No.(VC)': None, 'CADICS No.(PT1)': None, 'CADICS No.(PT2)': None,
                    'CADICS No.(PFC)': None,
                    'Name Table(VC)': None, 'Name Table(PT1)': None, 'Name Table(PT2)': None,
                    'Name Table(PFC)': None,
                    '配車使用欄(VC)': None, '配車使用欄(PT1)': None, '配車使用欄(PT2)': None,
                    '配車使用欄(PFC)': None,
                    'Table1-market(VC)': None, 'Table1-market(PT1)': None, 'Table1-market(PT2)': None,
                    'Table1-market(PFC)': None, 'WTC(VC)': None, 'WTC(PT1)': None, 'WTC(PT2)': None,
                    'WTC(PFC)': None}
        for item in list_sheets:

            if item not in all_sheets:
                dict_sub['file_name'] = file_name
                dict_sub[item + '_'] = '✕'
            else:
                df = pd.read_excel(file_path, sheet_name=item, header=None)
                if df.iloc[29, 0] != 'CADICS No.':
                    dict_sub['file_name'] = file_name
                    dict_sub[f'CADICS No.({item[3:]})'] = '✕'
                if df.iloc[0, 13]:
                    dict_sub['file_name'] = file_name
                    dict_sub[f'Name Table({item[3:]})'] = '✕'
                if df.iloc[1, 13] != '配車使用欄':
                    dict_sub['file_name'] = file_name
                    dict_sub[f'配車使用欄({item[3:]})'] = '✕'
                matching_columns_zone = df.columns[df.iloc[1].apply(lambda x: str(x).lower()) == '配車使用欄'].tolist()
                if len(matching_columns_zone) == 1:
                    continue
                if len(matching_columns_zone) > 1:
                    for id_table in matching_columns_zone:
                        if id_table == 13:
                            continue
                        else:
                            if not df.iloc[0, 14] and df.iloc[0, id_table + 1]:
                                dict_sub['file_name'] = file_name
                                dict_sub[f'Table1-market({item[3:]})'] = '✕'
                            for id_x_table_1 in range(15, id_table):
                                if df.iloc[24, id_x_table_1] == 0:
                                    continue
                                else:
                                    if df.iloc[20, id_x_table_1] not in ['W', 'T', 'C']:
                                        dict_sub['file_name'] = file_name
                                        dict_sub[f'WTC({item[3:]})'] = '✕'
                                    # if '':
        if list(set(dict_sub.values())) != [None]:
            list_end.append(dict_sub)
    df = pd.DataFrame(list_end)
    return df


def check_karen_4(karen4_files, karen4_path):
    list_sheets = ['関連表VC', '関連表PFC']
    list_end = []
    for file_name in karen4_files:
        file_path = karen4_path + '/' + file_name
        all_sheets = pd.ExcelFile(file_path).sheet_names
        for item in list_sheets:
            dict_sub = {'file_name': None, '関連表VC': None, '関連表PFC': None, 'CADICS No.(VC)': None,
                        '実験部品(VC)': None, '特性管理部品(VC)': None, 'CADICS No.(PFC)': None, '実験部品(PFC)': None,
                        '特性管理部品(PFC)': None}
            if item not in all_sheets:
                dict_sub['file_name'] = file_name
                dict_sub[item] = '✕'
            else:
                df = pd.read_excel(file_path, sheet_name=item, header=None)

                if df.iloc[18, 0] != 'CADICS No.':
                    dict_sub['file_name'] = file_name
                    dict_sub[f'CADICS No.({item[3:]})'] = '✕'
                if df.iloc[0, 12] != '実験部品':
                    dict_sub['file_name'] = file_name
                    dict_sub[f'実験部品({item[3:]})'] = '✕'
                matching_columns_hb = df.columns[df.iloc[0].apply(lambda x: str(x).lower()) == '特性管理部品'].tolist()
                if len(matching_columns_hb) == 0:
                    dict_sub['file_name'] = file_name
                    dict_sub[f'特性管理部品({item[3:]})'] = '✕'
            if list(set(dict_sub.values())) != [None]:
                list_end.append(dict_sub)
    df = pd.DataFrame(list_end)
    return df


#a=check_document(r"C:\XQZ_No_1\data\WZ1J", r"C:\XQZ_No_1\data\WZ1J\仕様表_WZ1J.xlsx",r"C:\XQZ_No_1\output\WZ1J_EV_US_CASE1.5")